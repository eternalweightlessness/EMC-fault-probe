from __future__ import annotations

from pathlib import Path

from emc_core.domain.fault_case import FaultCase
from openpyxl import load_workbook

from integrations.exporters.excel import ExcelCaseExporter


def test_excel_export_contains_headers_and_cases(tmp_path: Path) -> None:
    case = FaultCase(
        object_name="设备",
        phenomenon="辐射发射超标",
        cause="屏蔽问题",
        solution="改善屏蔽",
        severity="严重",
        frequency="偶发",
    )

    destination = ExcelCaseExporter().export([case], tmp_path / "result")

    assert destination.suffix == ".xlsx"
    workbook = load_workbook(destination)
    sheet = workbook["EMC 故障案例"]
    assert [cell.value for cell in sheet[1]] == list(FaultCase.FIELD_MAP)
    assert sheet.cell(row=2, column=2).value == "辐射发射超标"
    # Windows 会阻止删除仍被占用的 xlsx；显式 close 比等待垃圾回收更可靠。
    workbook.close()
